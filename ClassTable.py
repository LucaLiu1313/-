import pandas as pd
import openpyxl
import os
from openpyxl import load_workbook
from tkinter import Tk, Label, Entry, Button, filedialog, messagebox

# 创建 GUI 界面
root = Tk()
root.title("课表处理")
root.geometry("500x600")


# 选择 Excel 文件
def select_file():
    global file_path
    file_path = filedialog.askopenfilename(title="请选择 Excel 文件", filetypes=[("Excel 文件", "*.xlsx;*.xls")])
    if file_path:
        file_label.config(text=f"已选择: {os.path.basename(file_path)}")


# 选择保存 Excel 文件
def save_file():
    global excel_path
    excel_path = filedialog.asksaveasfilename(
        title="请选择保存 Excel 文件",
        defaultextension=".xlsx",
        filetypes=[("Excel 文件", "*.xlsx")])
    if excel_path:
        save_label.config(text=f"保存路径: {os.path.basename(excel_path)}")


# 处理 Excel 逻辑
def process_excel():
    name = name_entry.get().strip()  # 获取字典名称
    column_num = column_entry.get().strip()  # 获取列号

    if not name:
        messagebox.showwarning("警告", "请输入字典名称！")
        return
    if not file_path:
        messagebox.showwarning("警告", "请先选择 Excel 文件！")
        return
    if not excel_path:
        messagebox.showwarning("警告", "请选择 Excel 保存路径！")
        return
    if not column_num.isdigit():
        messagebox.showwarning("警告", "请输入正确的列号（数字）！")
        return

    column_num = int(column_num)

    # 创建字典
    globals()[name] = {}
    时间段 = []
    班级 = []

    # 读取 Excel 数据
    df = pd.read_excel(file_path)
    print(f"成功打开文件: {file_path}")
    print(df.head())

    # 遍历 DataFrame 查找 "语文战"
    for row_index, row in df.iloc[4:].iterrows():
        for col_index, col in enumerate(df.columns[1:], start=1):
            value = row[col]
            first_col_value = df.iloc[row_index, 0]  # 该行的第一列（时间段）
            first_row_value = df.iloc[0, col_index]  # 该列的第一行（班级）

            if value == name:
                时间段.append(first_col_value)
                班级.append(first_row_value)
                globals()[name][first_col_value] = first_row_value  # 确保字典有数据

    # 构建 DataFrame
    data = {"时间段": 时间段, "班级": 班级}
    df1 = pd.DataFrame(data)

    # 确保 Excel 文件存在
    if not os.path.exists(excel_path):
        df1.to_excel(excel_path, index=False, engine="openpyxl")  # 先创建文件

    print(df1)

    # 载入 Excel
    wb = load_workbook(excel_path, read_only=False, keep_vba=True)
    ws = wb.active

    # 遍历 "语文战" 数据并写入 Excel
    # 遍历 "语文战" 数据并写入 Excel
    for key, value in globals()[name].items():
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if str(cell.value).strip() == key:  # 确保 key 匹配
                    row_index = cell.row  # 获取匹配的行索引
                    target_cell = ws.cell(row=row_index, column=column_num)  # 目标单元格

                    if target_cell.value:  # 🔥 **检查单元格是否已填充**
                        messagebox.showerror("错误", f"单元格 (行 {row_index}, 列 {column_num}) 已有内容，无法覆盖！")
                        print(f"❌ 跳过写入: (行 {row_index}, 列 {column_num})，已有数据: {target_cell.value}")
                        continue  # 跳过写入

                    # 写入新数据
                    target_cell.value = value
                    print(f"✅ 成功写入: {key} -> 行 {row_index}, 班级: {value}")

    # 保存 Excel
    wb.save(excel_path)
    wb.close()

    # 处理完成提示
    messagebox.showinfo("成功", "Excel 数据处理完成！")


# GUI 组件
Label(root, text="请输入老师称呼:").pack(pady=5)
name_entry = Entry(root)
name_entry.pack(pady=5)

Label(root, text="请选择 Excel 文件:").pack(pady=5)
file_label = Label(root, text="未选择文件", fg="red")
file_label.pack(pady=5)
Button(root, text="选择文件", command=select_file).pack(pady=5)

Label(root, text="请选择 Excel 保存路径:").pack(pady=5)
save_label = Label(root, text="未选择保存路径", fg="red")
save_label.pack(pady=5)
Button(root, text="选择保存位置", command=save_file).pack(pady=5)

Label(root, text="请输入列号（数字）:").pack(pady=5)
column_entry = Entry(root)
column_entry.pack(pady=5)

Button(root, text="开始处理", command=process_excel).pack(pady=20)

# 运行 GUI 主循环
root.mainloop()
